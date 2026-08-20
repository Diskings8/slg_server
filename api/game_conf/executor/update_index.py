# -*- coding: utf-8 -*-
"""更新 Index.xlsx，添加新增的 hero 相关表。"""

from openpyxl import load_workbook
from pathlib import Path


def main():
    index_path = Path(__file__).parent.parent / "excel" / "Index.xlsx"

    # 加载现有的 Index.xlsx
    wb = load_workbook(index_path)
    ws = wb.active

    # 新增的表注册（插入到 hero_attr 之后）
    new_tables = [
        ('数据表', 'hero_base', 'hero_base.xlsx', None, '英雄基础信息'),
        ('数据表', 'hero_method', 'hero_method.xlsx', None, '战法配置'),
        ('数据表', 'hero_method_slot', 'hero_method_slot.xlsx', None, '英雄战法槽位'),
        ('数据表', 'hero_policy', 'hero_policy.xlsx', None, '政策配置'),
        ('数据表', 'hero_group', 'hero_group.xlsx', None, '羁绊配置'),
    ]

    # 找到 hero_attr 的位置
    hero_attr_row = None
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        if row[1].value == 'hero_attr':
            hero_attr_row = idx
            break

    if hero_attr_row is None:
        print("未找到 hero_attr 表")
        return

    print(f"在第 {hero_attr_row} 行之后插入 5 张新表")

    # 在 hero_attr 之后插入新行
    for i, new_table in enumerate(new_tables, start=1):
        insert_row = hero_attr_row + i
        ws.insert_rows(insert_row)
        ws.cell(insert_row, 1, new_table[0])
        ws.cell(insert_row, 2, new_table[1])
        ws.cell(insert_row, 3, new_table[2])
        ws.cell(insert_row, 4, new_table[3])
        ws.cell(insert_row, 5, new_table[4])

    # 保存
    wb.save(index_path)
    print(f"已更新: {index_path}")


if __name__ == '__main__':
    main()
