# -*- coding: utf-8 -*-
"""更新 game_attribute.xlsx，添加新增表的字段定义。

从 gameconfig_schema.py 中读取表定义，自动生成 game_attribute.xlsx 行。
"""

from openpyxl import load_workbook
from pathlib import Path
import sys

# 导入 schema
sys.path.insert(0, str(Path(__file__).parent))
from gameconfig_schema import TABLES, Field


def find_table_in_schema(table_name):
    """从 TABLES 中查找表定义。"""
    for table in TABLES:
        if table['name'] == table_name:
            return table
    return None


def generate_attribute_rows(table_def):
    """根据表定义生成 game_attribute.xlsx 的行数据。"""
    rows = []
    table_name = table_def['name']

    for field in table_def['fields']:
        row = [
            '表头',                          # 类别
            table_name,                      # 表名
            field.label,                     # 标识名（中文）
            field.name,                      # 字段名（英文）
            field.type,                      # 字段类型
            '|' if field.array else None,   # 数组切割符
            None,                            # 值
            '1' if field.index else None,   # 索引
            None,                            # 其他
            field.comment,                   # 备注
        ]
        rows.append(row)

    return rows


def main():
    excel_path = Path(__file__).parent.parent / "excel" / "game_attribute.xlsx"

    # 新增的表
    new_tables = ['hero_base', 'hero_method', 'hero_method_slot', 'hero_policy', 'hero_group']

    # 加载现有的 game_attribute.xlsx
    wb = load_workbook(excel_path)
    ws = wb.active

    # 找到最后一行
    last_row = ws.max_row
    print(f"当前 game_attribute.xlsx 有 {last_row} 行")

    # 为每个新表生成字段定义
    total_added = 0
    for table_name in new_tables:
        table_def = find_table_in_schema(table_name)
        if not table_def:
            print(f"警告: 未找到表 {table_name} 的 schema 定义")
            continue

        rows = generate_attribute_rows(table_def)
        print(f"为表 {table_name} 生成 {len(rows)} 行字段定义")

        # 追加到 worksheet
        for row in rows:
            ws.append(row)
            total_added += 1

    # 保存
    wb.save(excel_path)
    print(f"\n已更新: {excel_path}")
    print(f"总共添加了 {total_added} 行字段定义")


if __name__ == '__main__':
    main()
